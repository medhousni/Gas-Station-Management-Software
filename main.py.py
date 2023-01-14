import tkinter
from tkinter import *
from tkcalendar import DateEntry
from PIL import ImageTk, Image
from tkinter import ttk
from tkinter import messagebox
import os
import openpyxl
from tkinter.simpledialog import askstring


# Login page
login_page = tkinter.Tk()
login_page.title("Log in")
login_page.geometry("925x500+300+200")
login_page.configure(bg="white")
login_page.resizable(False, False)


def login():
    username = "admin"
    password = "1234"
    if username == user.get() and password == pw.get():
        login_page.destroy()
        window = tkinter.Tk()  # parent window
        window.title("Gas Station Management System")
        window.geometry("925x500")
        window.resizable(False, False)
        window.configure(bg="white")

        img = tkinter.PhotoImage(file="C:\\Users\\FINAL\\Downloads\\925.png")
        label = Label(window, image=img, width=925, height=500)
        label.pack()
        label.img = img

        Button(window, width=25, pady=7, text="Gas Management", bg="DarkOrange2", fg="white", border=0,
               command=gas_management).place(x=35, y=350)
        Button(window, width=25, pady=7, text="Inventory Management", bg="DarkOrange2", fg="white", border=0,
               command=inventory_management).place(x=365, y=350)
        Button(window, width=25, pady=7, text="Human Resources", bg="DarkOrange2", fg="white", border=0,
               command=hr).place(x=710, y=350)
    else:
        messagebox.showerror("Error", "Wrong credentials")


logo = tkinter.PhotoImage(file="C:\\Users\\FINAL\\Downloads\\total.png")
Label(login_page, image=logo, bg="white").place(x=50, y=85)

# frame containing the sign in form
frame1 = Frame(login_page, width=350, height=350, bg="white")
frame1.place(x=500, y=70)

heading = Label(frame1, text="Sign in", fg="DarkOrange2", bg="white", font=("Microsoft YaHei UI light", 23, "bold"))
heading.place(x=145, y=5)


# when selecting the username box remove the text "username"
def click_on(i):
    user.delete(0, "end")


# when leaving the username box print the text "username"
def click_off(i):
    name = user.get()
    if name == "":
        user.insert(0, "Username")


#   username entry form
user = Entry(frame1, width=25, fg="black", border=0, bg="white", font=("Microsoft YaHei UI light", 11))
user.place(x=70, y=80)
user.insert(0, "Username")
user.bind("<FocusIn>", click_on)
user.bind("<FocusOut>", click_off)

Frame(frame1, width=295, height=2, bg="black").place(x=60, y=105)


# when selecting the pw box remove the text "password"
def click_on(i):
    pw.delete(0, "end")


# when leaving the pw box print the text "password"
def click_off(i):
    name = pw.get()
    if name == "":
        pw.insert(0, "Password")


#  password entry form
pw = Entry(frame1, width=25, fg="black", border=0, bg="white", font=("Microsoft YaHei UI light", 11))
pw.place(x=70, y=150)
pw.insert(0, "Password")
pw.bind("<FocusIn>", click_on)
pw.bind("<FocusOut>", click_off)



Frame(frame1, width=295, height=2, bg="black").place(x=60, y=177)

#########################

Signin_button = Button(frame1, width=35, pady=7, text="Sign in", bg="DarkOrange2", fg="white", border=0, command=login)
Signin_button.place(x=78, y=235)

Exit_button = Button(login_page, width=5, pady=7, text="Exit", bg="DarkOrange2", fg="white", border=0,
                     command=login_page.destroy)
Exit_button.place(x=850, y=440)


def gas_management():
    name = askstring('File', 'Save as:')
    window = tkinter.Tk()  # parent window
    window.resizable(False, False)
    window.title("Gas Purchase Management System")

    frame = tkinter.Frame(window)
    frame.pack()

    DateOfpurchase_label = tkinter.Label(frame, text="Date of purchase")
    DateOfpurchase_label.grid(row=0, column=0)
    DateOfpurchase_entry = DateEntry(frame, selectmode='day')
    DateOfpurchase_entry.grid(row=1, column=0)

    Gas_type_label = tkinter.Label(frame, text="Gas Type")
    Gas_type_label.grid(row=0, column=1)
    Gas_type_box = ttk.Combobox(frame, values=["", "Gasoil", "S.S.P"])
    Gas_type_box.grid(row=1, column=1)

    Price_label = tkinter.Label(frame, text="Price")
    Price_label.grid(row=0, column=2)
    Price_entry = tkinter.Entry(frame)
    Price_entry.grid(row=1, column=2)

    GasStation_Location = tkinter.Label(frame, text="Location")
    GasStation_Location.grid(row=2, column=1)
    Location = ttk.Combobox(frame, values=["", "Location 1", "Location 2"])
    Location.grid(row=3, column=1)

    # Result Display

    columns = ("Date", "Gas type", "Price")
    trv = ttk.Treeview(frame, columns=columns, show="headings")
    trv.grid(row=4, column=0, columnspan=3, padx=20, pady=10)
    trv.heading("Date", text="Date")
    trv.heading("Gas type", text="Gas type")
    trv.heading("Price", text="Price")
    trv.tag_configure('oddrow', background="white")
    trv.tag_configure('oddrow', background="lightblue")

    def clear_data():
        DateOfpurchase_entry.delete(0, tkinter.END)
        Gas_type_box.delete(0, tkinter.END)
        Price_entry.delete(0, tkinter.END)

    global count
    count = 0
    def enter_data():
        global count
        date = DateOfpurchase_entry.get()
        gas = Gas_type_box.get()
        cost = float(Price_entry.get())
        table_items = [date, gas, cost]
        if count % 2 == 0:
            trv.insert('', 0, values=table_items, tags=("evenrow"))
        else:
            trv.insert('', 0, values=table_items, tags=("oddrow"))
        count +=1

        clear_data()
        filepath = "C:\\Users\\FINAL\\Desktop\\IMS\\" + name + ".xlsx"
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["Date", "Gas type", "Cost"]
            sheet.append(heading)
            workbook.save(filepath)
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        sheet.append([date, gas, cost])
        workbook.save(filepath)

    data_totable = tkinter.Button(frame, text="Enter Data", command=enter_data)
    data_totable.grid(row=3, column=2)


def inventory_management():
    name = askstring('File', 'Save as:')
    window = tkinter.Tk()  # parent window
    window.resizable(False, False)
    window.title("Inventory Management System")

    frame = tkinter.Frame(window)
    frame.pack()

    DateOfpurchase_label = tkinter.Label(frame, text="Date of purchase")
    DateOfpurchase_label.grid(row=0, column=0)
    DateOfpurchase_entry = DateEntry(frame, selectmode='day')
    DateOfpurchase_entry.grid(row=1, column=0)

    Product_label = tkinter.Label(frame, text="Product")
    Product_label.grid(row=0, column=1)
    Product_entry = tkinter.Entry(frame)
    Product_entry.grid(row=1, column=1)

    Units_label = tkinter.Label(frame, text="Units")
    Units_label.grid(row=0, column=2)
    Units_entry = tkinter.Entry(frame)
    Units_entry.grid(row=1, column=2)

    Price_label= tkinter.Label(frame, text="Price per unit")
    Price_label.grid(row=2, column=1)
    Price_entry = tkinter.Entry(frame)
    Price_entry.grid(row=3, column=1)

    for widget in frame.winfo_children():       #padding between input boxes
        widget.grid_configure(padx=10, pady=5)

    # Result Display

    columns = ("Date", "Product", "Units", "PPU", "Total")
    trv = ttk.Treeview(frame, columns=columns, show="headings")
    trv.grid(row=4, column=0, columnspan=3, padx=20, pady=10)
    trv.heading("Date", text="Date")
    trv.heading("Product", text="Product")
    trv.heading("Units", text="Units")
    trv.heading("PPU", text="PPU")
    trv.heading("Total", text="Total")
    trv.tag_configure('oddrow', background="white")
    trv.tag_configure('oddrow', background="lightblue")

    def clear_data():
        DateOfpurchase_entry.delete(0, tkinter.END)
        Product_entry.delete(0, tkinter.END)
        Units_entry.delete(0, tkinter.END)
        Price_entry.delete(0, tkinter.END)


    global count
    count = 0

    def enter_data():
        global count
        date = DateOfpurchase_entry.get()
        product = Product_entry.get()
        units = float(Units_entry.get())
        ppu = float(Price_entry.get())
        total = units * ppu
        table_items = [date, product, units, ppu, total]
        if count % 2 == 0:
            trv.insert('', 0, values=table_items, tags=("evenrow"))
        else:
            trv.insert('', 0, values=table_items, tags=("oddrow"))
        count += 1

        clear_data()
        filepath = "C:\\Users\\FINAL\\Desktop\\IMS\\" + name + ".xlsx"
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["Date", "Products", "Units", "PPU", "Total"]
            sheet.append(heading)
            workbook.save(filepath)
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        sheet.append([date, product, units, ppu, total])
        workbook.save(filepath)

    data_totable = tkinter.Button(frame, text="Enter Data", command=enter_data)
    data_totable.grid(row=3, column=2)
    # #clearTable = tkinter.Button(frame, text="Clear table", command=clear_table)
    # #clearTable.grid(row=5, column=0, padx=20, pady=10)




def hr():
    name = askstring('File', 'Save as:')
    window = tkinter.Tk()  # parent window
    window.resizable(False, False)
    window.title("Human Resources Management")

    frame = tkinter.Frame(window)
    frame.pack()


    # Saving user info
    info_frame = tkinter.LabelFrame(frame, text="Staff Information")
    info_frame.grid(row=0, column=0, padx=15, pady=15)

    name_label = tkinter.Label(info_frame, text="First Name")
    name_label.grid(row=0, column=0)
    last_name_label = tkinter.Label(info_frame, text="Last Name")
    last_name_label.grid(row=0, column=1)
    Id = tkinter.Label(info_frame, text="Identification Number")
    Id.grid(row=0, column=2)
    salary = tkinter.Label(info_frame, text="Salary")
    salary.grid(row=0, column=3)

    name_entry = tkinter.Entry(info_frame)
    last_name_entry = tkinter.Entry(info_frame)
    Id_entry = tkinter.Entry(info_frame)
    salary_entry = tkinter.Entry(info_frame)
    name_entry.grid(row=1, column=0)
    last_name_entry.grid(row=1, column=1)
    Id_entry.grid(row=1, column=2)
    salary_entry.grid(row=1, column=3)

    for widget in info_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    def clear_data():
        name_entry.delete(0, tkinter.END)
        last_name_entry.delete(0, tkinter.END)
        Id_entry.delete(0, tkinter.END)
        salary_entry.delete(0, tkinter.END)

    global count
    count = 0
    def enter_data():
        global count
        firstName = name_entry.get()
        lastName = last_name_entry.get()
        id = Id_entry.get()
        pay = float(salary_entry.get())

        table_items = [firstName, lastName, id, pay]
        if count % 2 == 0:
            trv.insert('', 0, values=table_items, tags=("evenrow"))
        else:
            trv.insert('', 0, values=table_items, tags=("oddrow"))
        count +=1

        clear_data()
        filepath = "C:\\Users\\FINAL\\Desktop\\IMS\\" + name + ".xlsx"
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["First Name", "Last Name", "Identification Number", "Salary"]
            sheet.append(heading)
            workbook.save(filepath)
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        sheet.append([firstName, lastName, id, pay])
        workbook.save(filepath)

    columns = ("First Name", "Last Name", "Identification Number","Salary")
    trv = ttk.Treeview(frame, columns=columns, show="headings")
    trv.grid(row=4, column=0, columnspan=3, padx=20, pady=10)
    trv.heading("First Name", text="First Name")
    trv.heading("Last Name", text="Last Name")
    trv.heading("Identification Number", text="Identification Number")
    trv.heading("Salary", text="Salary")
    trv.tag_configure('oddrow', background="white")
    trv.tag_configure('oddrow', background="lightblue")

    data_totable = tkinter.Button(frame, text="Enter Data", command=enter_data)
    data_totable.grid(row=3, column=2)



mainloop()
